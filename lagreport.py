#!/usr/bin/env python
# lag.py [cuttoffDate] [startDate]

import os
import sys
import datetime
import workdays
import pymssql
from openpyxl import Workbook
from openpyxl.styles import Font, Fill,  PatternFill,  Alignment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from collections import OrderedDict
from dotenv import load_dotenv

load_dotenv()

# Holidays Off  from https://www.timeanddate.com/holidays/us/2021
holidays = [(lambda x:  datetime.datetime.strptime(x ,"%Y-%m-%d").date()  ) (x) for x in [
    "2020-01-01","2020-05-25","2020-07-03","2020-09-07","2020-11-26","2020-11-27","2020-12-25",
    "2020-12-31","2021-05-31","2021-07-05","2021-09-06","2021-11-25","2021-11-26","2021-12-25",
    "2021-12-31","2022-05-30","2022-07-04","2022-09-05","2022-11-24","2022-11-25","2022-12-24",
    "2023-01-02","2023-05-29","2023-07-04","2023-09-04","2023-11-22","2023-11-23","2023-12-26",
    "2024-01-01","2024-05-27","2024-07-04","2024-09-02","2024-11-28","2024-11-29","2024-12-25",
    "2025-01-01","2025-05-26","2025-07-04","2025-09-01","2025-11-27","2025-11-28","2025-12-25",
    "2026-01-01","2026-05-25","2026-07-03","2026-09-07","2026-11-26","2026-11-27","2026-12-25",
    "2027-01-01","2027-05-31","2027-07-05","2027-09-06","2027-11-25","2027-11-26","2027-12-24",
    "2028-01-01","2028-05-29","2028-07-04","2028-09-04","2028-11-23","2028-11-24","2028-12-25",
    "2029-01-01","2029-05-28","2029-07-04","2029-09-03","2029-11-22","2029-11-23","2029-12-25",
]]

def outputcsv(boat):
    if boat["canvasStart"] == "":
        lag1 = "N/A"
        lag2 = max(workdays.networkdays(datetime.datetime.strptime(boat["fabEnd"],"%Y-%m-%d").date(), \
                   datetime.datetime.strptime(boat["paintStart"],"%Y-%m-%d").date(), holidays) -2, 0)
    else:
        lag1 = max(workdays.networkdays(datetime.datetime.strptime(boat["fabEnd"],"%Y-%m-%d").date(), \
                   datetime.datetime.strptime(boat["canvasStart"],"%Y-%m-%d").date(), holidays) -2, 0)
        lag2 = max(workdays.networkdays(datetime.datetime.strptime(boat["canvasEnd"],"%Y-%m-%d").date(), \
                   datetime.datetime.strptime(boat["paintStart"],"%Y-%m-%d").date(), holidays) -2, 0)
    lag3 = max(workdays.networkdays(datetime.datetime.strptime(boat["paintEnd"],"%Y-%m-%d").date(), \
               datetime.datetime.strptime(boat["outfitStart"],"%Y-%m-%d").date(), holidays) -2, 0)
    print('"%s","%s","%s","%s","%s","%s","%s","%s","%s","%s","%s","%s"' % \
          (boat["job"], boat["fabStart"], boat["fabEnd"], boat["canvasStart"], lag1, boat["canvasEnd"], \
           boat["paintStart"], lag2, boat["paintEnd"], boat["outfitStart"], lag3, boat["outfitEnd"]))



def writeexcel(box, cutoffDate):
    wb = Workbook()
    ws = wb.active
    bld = Font(bold=True)
    ws.title = "Lag Report"

    # '"Boat","Fab Start","Fab End","Canvas Start","Lag 1","Canvas End","Paint Start","Lag 2","Paint End","Outfit Start","Lag 3","Outfit End"'

    props = [ ['A', 'Boat', 10.2602040816327], ['B', 'Firstday\nFab', 12.6887755102041], ['C', 'Lastday\nFab', 12.6887755102041], \
              ['D', 'Firstday\nCanvas', 12.6887755102041], ['E', 'Lag', 10.2602040816327], ['F', 'Lastday\nCanvas', 12.6887755102041], \
              ['G', 'Firstday\nPaint', 12.6887755102041], ['H', 'Lag', 10.2602040816327], ['I', 'Lastday\nPaint', 12.6887755102041], \
              ['J', 'Firstday\nOutfitting', 12.6887755102041], ['K', 'Lag', 10.2602040816327], ['L', 'Lastday\nOutfitting', 12.6887755102041] ]

    for col, text, width in props:
        ws['%s1'%col] = text
        ws.column_dimensions[col].width = width
        currentCell = ws.cell(1, ord(col)-64)
        currentCell.alignment = Alignment(horizontal='center')
        currentCell.font = bld

    # set header row height
    ws.row_dimensions[1].height = 35.05

    row_index = 2
    for boat in sorted(box, key=lambda k: k['outfitStart']):
        if boat["outfitEnd"] != "":
            column_index = 1

            if datetime.datetime.strptime(boat["outfitStart"] ,"%Y-%m-%d").date() > cutoffDate:
                if boat["canvasStart"] == "":
                    # no canvas stage only output for paint and outfitting
                    lag1 = ""
                    lag2 = max(workdays.networkdays(datetime.datetime.strptime(boat["fabEnd"],"%Y-%m-%d").date(), \
                               datetime.datetime.strptime(boat["paintStart"],"%Y-%m-%d").date(), holidays) -2, 0)
                else:
                    lag1 = max(workdays.networkdays(datetime.datetime.strptime(boat["fabEnd"],"%Y-%m-%d").date(), \
                               datetime.datetime.strptime(boat["canvasStart"],"%Y-%m-%d").date(), holidays) -2, 0)
                    lag2 = max(workdays.networkdays(datetime.datetime.strptime(boat["canvasEnd"],"%Y-%m-%d").date(), \
                               datetime.datetime.strptime(boat["paintStart"],"%Y-%m-%d").date(), holidays) -2, 0)
                    lag3 = max(workdays.networkdays(datetime.datetime.strptime(boat["paintEnd"],"%Y-%m-%d").date(), \
                               datetime.datetime.strptime(boat["outfitStart"],"%Y-%m-%d").date(), holidays) -2, 0)

                ws["A%s"%(row_index)] = boat["job"]
                ws["B%s"%(row_index)] = boat["fabStart"]
                ws["C%s"%(row_index)] = boat["fabEnd"]
                ws["D%s"%(row_index)] = boat["canvasStart"]
                ws["E%s"%(row_index)] = lag1
                ws["F%s"%(row_index)] = boat["canvasEnd"]
                ws["G%s"%(row_index)] = boat["paintStart"]
                ws["H%s"%(row_index)] = lag2
                ws["I%s"%(row_index)] = boat["paintEnd"]
                ws["J%s"%(row_index)] = boat["outfitStart"]
                ws["K%s"%(row_index)] = lag3
                ws["L%s"%(row_index)] = boat["outfitEnd"]

                row_index += 1

    greyFill = PatternFill(start_color='FFCCCCCC', end_color='FFCCCCCC', fill_type='solid')
    ws.conditional_formatting.add('A2:L%s'%(row_index-1), FormulaRule(formula=['ISEVEN(ROW())'], stopIfTrue=True, fill=greyFill))
    wb.save('/tmp/LagReport-%s.xlsx'%(datetime.date.today()))



def lagReport():
    onlyBoat = ""
    target_date = datetime.date.today()
    # target_date = datetime.datetime.strptime('2018-09-10', '%Y-%m-%d').date()
    startDate = target_date -  datetime.timedelta(weeks=16)
    cutoffDate = target_date - datetime.timedelta(days=61)

    if len(sys.argv) > 1:   # set cutoff date
        cutoffDate =  datetime.datetime.strptime(sys.argv[1] ,"%Y-%m-%d").date()

    if len(sys.argv) > 2:   # set start date
        startDate =  datetime.datetime.strptime(sys.argv[2] ,"%Y-%m-%d").date()

    SQL = """
    SELECT job.jobname, twp.workingpunch_ts,
           CASE twp.department_id
                WHEN 207 THEN 'Fab'
                WHEN 221 THEN 'Canvas'
                WHEN 225 THEN 'Canvas'
                WHEN 213 THEN 'Paint'
                WHEN 218 THEN 'Paint'
           ELSE '' END AS dept
      FROM job
INNER JOIN timeWorkingPunch twp ON job.job_id = twp.job_id
     WHERE (   JobName LIKE '% 717' OR  JobName LIKE '% 718'
            OR JobName LIKE '% 818' OR  JobName LIKE '% 819'
            OR JobName LIKE '% 919' OR  JobName LIKE '% 920'
            OR JobName LIKE '% 020' OR  JobName LIKE '% 021'
            OR JobName LIKE '% 121' OR  JobName LIKE '% 122'
            OR JobName LIKE '% 222' OR  JobName LIKE '% 223'
            OR JobName LIKE '% 323' OR  JobName LIKE '% 324'
            OR JobName LIKE '% 424' OR  JobName LIKE '% 425' )
        AND start_ts > '""" + startDate.strftime("%Y-%m-%d") + """ 23:00:00'
        ORDER BY job.jobname, twp.workingpunch_ts;"""

    boat = { "job": "", "fabStart": "", "fabEnd": "", "canvasStart": "", "canvasEnd": "", \
             "paintStart": "", "paintEnd": "", "outfitStart": "", "outfitEnd": "" }

    conn = pymssql.connect(os.getenv('DB_HOST'),os.getenv('DB_USER'),os.getenv('DB_PASSWORD'),os.getenv('DB_DATABASE'))
    cursor = conn.cursor()
    cursor.execute(SQL)
    cur = conn.cursor(as_dict=True)
    box = []
    for row in cur:
        punch = row["workingpunch_ts"]
        punchday = punch.strftime("%Y-%m-%d")
        dept = row["dept"]
        job = row["jobname"]

        if boat["job"] != job:
            if boat["job"] != "":
                box.append(boat)
            del boat
            boat = { "job": "", "fabStart": "", "fabEnd": "", "canvasStart": "", "canvasEnd": "", \
                     "paintStart": "", "paintEnd": "", "outfitStart": "", "outfitEnd": "" }
            boat["job"] = job
            mode = "None"


        # MODE NONE =============================================================
        # ignore any punches before boat is in fab
        # if dept != "Fab" and mode == "none":
        #    ignore

        # MODE FAB  =============================================================
        if dept == "Fab" and mode == "None":
            mode = "Fab"
            boat["fabStart"] = punchday
            boat["fabEnd"] = punchday
            # send email to James Green

        if dept == "Fab" and mode == "Fab":
            boat["fabEnd"] = punchday
            day = punchday

        # ignore days with both fab and canvas punches
        if dept == "Canvas" and mode == "Fab" and day != punchday:
            boat["canvasStart"] = punchday
            boat["canvasEnd"] = punchday
            day = punchday
            mode = "Canvas"

        # ignore days with both fab and paint punches
        if dept == "Paint" and mode == "Fab" and day != punchday:
            boat["paintStart"] = punchday
            boat["paintEnd"] = punchday
            day = punchday
            mode = "Paint"

        # MODE CANVAS ===========================================================
        if dept == "Canvas" and mode == "Canvas":
            boat["canvasEnd"] = punchday
            day = punchday

        # ignore canvas and days with both canvas and paint punches
        if dept == "Paint" and mode == "Canvas" and day != punchday:
            boat["paintStart"] = punchday
            boat["paintEnd"] = punchday
            day = punchday
            mode = "Paint"


        # MODE PAINT ============================================================
        if dept == "Paint" and mode == "Paint":
            boat["paintEnd"] = punchday

        # ignore paint and days with both paint and canvas punches
        # canvas punches should be for
        if dept == "Canvas" and mode == "Paint" and day != punchday:
            boat["outfitStart"] = punchday
            boat["outfitEnd"] = punchday
            day = punchday
            mode = "Outfit"

        # MODE OUTFITTING =======================================================
        if dept == "Canv" and mode == "Outfit":
            boat["outfitEnd"] = punchday

    # PRINT LAST BOOT OUTSIDE OF LOOP ===========================================
    #box.append(boat)

    if (False):
        print('"Boat","Fab Start","Fab End","Canvas Start","Lag 1","Canvas End","Paint Start","Lag 2","Paint End","Outfit Start","Lag 3","Outfit End"')
        for boat in sorted(box, key=lambda k: k['outfitStart']):
            if boat["outfitEnd"] != "":
                if datetime.datetime.strptime(boat["outfitStart"] ,"%Y-%m-%d").date() > cutoffDate:
                    outputcsv(boat)

    writeexcel(sorted(box, key=lambda k: k['outfitStart']),cutoffDate)




if __name__ == "__main__":
    lagReport()
